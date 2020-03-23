#------------------------------------------------------------------------------------------------------------------------#
#---------------------------------------------|MINERVA|------------------------------------------------------------------#
#------------------------------------------------------------------------------------------------------------------------#
#---------------MINERVA is a tool that can be used to find experts by utilizing LinkedIn---------------------------------#
#---------------MINERVA is used only as a productivity tool and any data retrieved is not sold---------------------------#
#-----------------------------------------------------------------------------Developer: George Koureas------------------#
#-----------------------------------------------------------------------------Date: March 2020---------------------------#
#-----------------------------------------------------------------------------All rights reserved------------------------#
#------------------------------------------------------------------------------------------------------------------------#

import time
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.expected_conditions import presence_of_element_located
from selenium.webdriver.support.expected_conditions import visibility_of_element_located
from bs4 import BeautifulSoup
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException



class LinkedInBot:

    def __init__(self, username, password):

        # Initializes Chromedriver
        self.browser = webdriver.Chrome('./chromedriver.exe')
        self.wb = load_workbook('./database.xlsx')

        #setting url(s), username and pass which is read from config.txt
        self.base_url = 'https://www.linkedin.com'
        self.login_url = self.base_url + '/login'
        #self.feed_url = self.base_url + '/feed'
        #---------------------------------------------MAKE IT FUNCTIONAL FOR MORE THAN ONE PAGE BY ADDING HERE
        self.search_people_url = self.base_url + '/search/results/people/?keywords={search}&page={page}'


    def nav(self, url):
        self.browser.get(url)
        time.sleep(2)

    
    def login(self, username, password):
        #Loging in
        self.nav(self.login_url)
        self.browser.find_element_by_id('username').send_keys(username)
        self.browser.find_element_by_id('password').send_keys(password)
        time.sleep(3)
        self.browser.find_element_by_xpath("//button[contains(text(), 'Sign in')]").click()
        time.sleep(3)
    


    def search_people(self, text, profvis, row_count, page_no, connect=False):
        #initializing Project Sheet
        #if not (self.wb.get_sheet('Project1')):
        if profvis == 2:
            if page_no < 3:
                page_no = page_no + 1
                profvis = 0

    
        if not ('Project1' in self.wb.sheetnames):
            project_sheet = self.wb.create_sheet('Project1')
            project_sheet.cell(row = 1, column = 1).value = 'Name'
            project_sheet.cell(row = 1, column = 2).value = 'Link'
            project_sheet.cell(row = 1, column = 3).value = 'Job Positions'
            project_sheet.cell(row = 1, column = 4).value = 'Total Score'
            self.wb.save('database.xlsx')
        
        project_sheet = self.wb.get_sheet_by_name('Project1')
        self.wb.Alignment = Alignment (wrap_text=True)
        

        #searches from homescreen
        actions = ActionChains(self.browser)
        self.nav(self.search_people_url.format(search= text, page= page_no))
        time.sleep(2)
        wait = WebDriverWait(self.browser, 10)
        profilesID = wait.until(presence_of_element_located((By.CLASS_NAME, "search-result__image ")))
        profilesID = self.browser.find_elements_by_class_name("search-result__image ")
        connect_btn = profilesID[4]
        #print(profilesID)
        actions.move_to_element(connect_btn).perform()
        profilesID = wait.until(presence_of_element_located((By.CLASS_NAME, "search-result__image ")))
        profilesID = self.browser.find_elements_by_class_name("search-result__image ")
        connect_btn = profilesID[profvis]
        self.browser.execute_script("arguments[0].click();", connect_btn)
        time.sleep(2)
        project_sheet.cell(row = row_count, column = 2).value = self.browser.current_url
        self.browser.maximize_window()
        self.browser.implicitly_wait(5)

        score = 0
        #multipl = 0
        element = wait.until(presence_of_element_located((By.XPATH, ".//div[@id = 'oc-background-section']")))
        element = self.browser.find_element_by_xpath(".//div[@id = 'oc-background-section']")
        actions = ActionChains(self.browser)
        actions.move_to_element(element).perform()
        time.sleep(1)
        project_sheet.cell(row = row_count, column = 4).value = 0
        project_sheet.cell(row = row_count, column = 3).value = ' '


#-------------------------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------------------------
#---------NUMBER OF COMPANIES IN HIS CV-----------------------------------------------------------------------
#-------------------------------------------------------------------------------------------------------------



        if len(self.browser.find_elements_by_xpath(".//*[@id='experience-section']/ul/li[1]")) != 0:
            #----HERE we find the number of companies the person has worked at
            number_of_companies = self.browser.find_elements_by_xpath(".//*[@id='experience-section']/ul/li")
            
            
            
#--------------------------------SEARCH KEYWORDS---------------------------------------------------------#            
            
            position_keywords = ['logistics', 'ecommerce', 'fullfilment', 'parcel delivery']
            #company_keywords = ['coolblue', 'bol.com', 'dhl']
            country_keywords = ['netherlands', 'benelux', 'amsterdam', 'utrecht', 'nl', 'nederland']

#--------------------------------------------------------------------------------------------------------#




            for i in range(1,(len(number_of_companies)+1)):
                #----HERE we check if the person has changed multiple positions in the SAME company
                metr_check = str(i)
                #----IF he/she has, we find the total duration of his work and the titles of the position
                #----Might need to add specific durations, afterwards





#------------------------------------------------------------------------------------------------------------------------#
#----------------------------SECTION-1   -> Person has MULTIPLE jobs in same company-------------------------------------#
#------------------------------------------------------------------------------------------------------------------------#



                
                
                #SENIOR/MANAGER -> +1
                #CONSULTANT -> +2
                #CHIEF OFFICER/DIRECTOR -> +3

                if len(self.browser.find_elements_by_xpath(".//*[@id='experience-section']/ul/li[{}]//ul".format(metr_check))) != 0:
                    #total_duration = self.browser.find_element_by_xpath(".//*[@id='experience-section']/ul/li[{}]//*[@class= 'display-flex justify-space-between full-width']//a//*[@class= 'pv-entity__company-details']/div[2]/h4".format(metr_check))
                    time.sleep(2)
                    #total_duration_text = total_duration.text
                    #duration = [int(s) for s in total_duration_text.split() if s.isdigit()]
                    positions_number = self.browser.find_elements_by_xpath(".//*[@id='experience-section']/ul/li[{}]//ul/li".format(metr_check))
                    
                    for j in range(1,(len(positions_number)+1)):
                        mult = 0
                        metr = str(j)

                        #----------------------JOB TITLE-----------------------#
                        job_title = self.browser.find_element_by_xpath(".//*[@id='experience-section']/ul/li[{company_number}]//ul/li[{position_number}]//h3//span[2]".format(company_number= metr_check, position_number= metr))
                        time.sleep(1)
                        title_text = job_title.text
                        project_sheet.cell(row = row_count, column = 3).value = project_sheet.cell(row = row_count, column = 3).value + '\n' + title_text
                        #search_term = text.split()


                        for keyword in position_keywords:
                            if (keyword in title_text.lower()):
                                mult = 1
                                if (('senior' in title_text.lower()) or ('manager' in title_text.lower())):
                                    mult = mult + 0.5
                                if (('consultant' in title_text.lower()) or ('director' in title_text.lower())):
                                    mult = mult + 1
                                if (('chief' in title_text.lower()) and ('officer' in title_text.lower())):
                                    mult = mult + 2

                        #--------------------COUNTRY----------------------------#
                        country = self.browser.find_elements_by_xpath(".//*[@id='experience-section']/ul/li[{company_number}]//ul/li[{position_number}]//h4[@class= 'pv-entity__location t-14 t-black--light t-normal block']//span[2]".format(company_number= metr_check, position_number= metr))
                        
                        if len(country) != 0:
                            country_text = country[0].text
                            for keyword in country_keywords:
                                if ((keyword in country_text.lower()) and (mult > 0)):
                                    mult = mult + 1


                        
                        #--------------------POSITION DURATION------------------#
                        position_duration = self.browser.find_element_by_xpath(".//*[@id='experience-section']/ul/li[{company_number}]//ul/li[{position_number}]//div[@class= 'display-flex']//h4[2]//span[2]".format(company_number= metr_check, position_number= metr))
                        position_duration_text = position_duration.text
                        duration_number = [int(s) for s in position_duration_text.split() if s.isdigit()]
                        duration_split = position_duration_text.split()
                        
                        
                        

                        #------------------------SCORE-------------------------#
                        if len(duration_number) == 2:
                            score = score + ((duration_number[1]*0.083) + duration_number[0])*mult
                        else:
                            if ((duration_split[1] == 'mos') or (duration_split[1] == 'mo')):
                                score = score + (duration_number[0]*0.083)*mult
                            elif (('less' in duration_split)):
                                score = score + 0.083*mult
                            else:
                                score = score + (duration_number[0]*mult)
                        #------------------------------------------------------------------------
                        print(title_text)
                        print(position_duration_text)
                        print(mult)
                        project_sheet.cell(row = row_count, column = 4).value = float(project_sheet.cell(row = row_count, column = 4).value) + score




#-----------------------------------------------------------------------------------------------------------------------#      
#-----------------------SECTION-2   -> Person has ONE job in same company-----------------------------------------------#
#-----------------------------------------------------------------------------------------------------------------------#


                elif len(self.browser.find_elements_by_xpath(".//*[@id='experience-section']/ul/li[{}]".format(metr_check))) != 0:
                    
                    mult = 0
                    #-------------------JOB TITLE---------------------------------------------------------#
                    job_title = self.browser.find_element_by_xpath(".//*[@id='experience-section']/ul/li[{}]//*[@class= 'display-flex justify-space-between full-width']//a//div[2]/h3".format(metr_check))
                    time.sleep(1)
                    title_text = job_title.text
                    project_sheet.cell(row = row_count, column = 3).value = project_sheet.cell(row = row_count, column = 3).value + '\n' + title_text
                    #print(title_text)

                    for keyword in position_keywords:
                            if (keyword in title_text.lower()):
                                mult = 1
                                if (('senior' in title_text.lower()) or ('manager' in title_text.lower())):
                                    mult = mult + 0.5
                                if (('consultant' in title_text.lower()) or ('director' in title_text.lower())):
                                    mult = mult + 1
                                if (('chief' in title_text.lower()) and ('officer' in title_text.lower())):
                                    mult = mult + 2




                    #--------------------COUNTRY----------------------------#
                    country = self.browser.find_elements_by_xpath(".//*[@id='experience-section']/ul/li[{}]//*[@class= 'display-flex justify-space-between full-width']//a//div[2]//*[@class = 'pv-entity__location t-14 t-black--light t-normal block']".format(metr_check))
                    if len(country) != 0:
                        country_text = country[0].text
                        for keyword in country_keywords:
                            if ((keyword in country_text.lower()) and (mult >0)):
                                mult = mult + 1





                    #-------------------POSITION DURATION-------------------#
                    if len(self.browser.find_elements_by_xpath(".//*[@id='experience-section']/ul/li[{}]//*[@class= 'display-flex justify-space-between full-width']//a//div[2]//*[@class = 't-14 t-black--light t-normal']".format(metr_check))) != 0:
                        experience_duration = self.browser.find_element_by_xpath(".//*[@id='experience-section']/ul/li[{}]//*[@class= 'display-flex justify-space-between full-width']//a//div[2]//*[@class = 't-14 t-black--light t-normal']".format(metr_check))
                        time.sleep(2)
                        duration_text = experience_duration.text
                        #duration = [int(s) for s in duration_text.split() if s.isdigit()]
                        print(title_text)
                        print(duration_text)
                        duration_number = [int(s) for s in duration_text.split() if s.isdigit()]
                        duration_split = duration_text.split()
                    
                    #---------------DUARATION SCORE---------------------------------------------------------------
                        if len(duration_number) == 2:
                            score = score + ((duration_number[1]*0.083) + duration_number[0])*mult
                        else:
                            if ((duration_split[3] == 'mos') or (duration_split[3] == 'mo')):
                                score = score + (duration_number[0]*0.083)*mult
                            elif (('less' in duration_split)):
                                score = score + 0.083*mult
                            else:
                                score = score + (duration_number[0]*mult)
                        #---------------------------------------------------------------------------------------------
                        print(mult)
                        project_sheet.cell(row = row_count, column = 4).value = float(project_sheet.cell(row = row_count, column = 4).value) + score
        
        
        
#-----------------------------------------------------------------------------------------------------------------------#        
#-----------------------------------------------EXTRACTION TO EXCEL SHEET-----------------------------------------------#
#-----------------------------------------------------------------------------------------------------------------------#









#------------------------------------------------------------------------------------------------------------------------#
#----------------------CHANGE NUMBER OF PROFILES CHECKED HERE------------------------------------------------------------#
#------------------------------------------------------------------------------------------------------------------------#

        self.wb.save('database.xlsx')
        if profvis <= 1:
            row_count = row_count + 1
            profvis = profvis + 1
            self.search_people(text, profvis, row_count, page_no)




        




        




# CHANGE THE SEARCH URL IN ORDER TO INCORPORATE SEARCH FILTERS SUCH AS COMPANY AND LOCATION




if __name__ == '__main__':

    file = open('config.txt')
    lines = file.readlines()
    username = lines[0]
    password = lines[1]
    #visitedProfiles = []
    #ProfilesQueued = []
    bot = LinkedInBot(username, password)
    
    search_text = 'logistics netherlands senior manager ' #The search text should always be in lowercase

    row_count = 2
    bot.login(username, password)
    bot.search_people(search_text, 0, row_count, 1)
    time.sleep(1)
    #bot.visit_profile()
    #bot.getprofileIDs(visitedProfiles, ProfilesQueued)



